import os
from typing import Optional, Union
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from fileserver.connection import get_mounted_path, mount_network
from datetime import datetime
from pathlib import Path


def read_excel(
    file_path: str,
    sheet_name: Optional[Union[str, int]] = 0,
    network_path: Optional[str] = None,
    auto_mount: bool = True
) -> pd.DataFrame:
    if auto_mount:
        try:
            mount_network(network_path)
        except Exception:
            pass
    
    full_path = get_mounted_path(file_path, network_path)
    
    if not os.path.exists(full_path):
        raise FileNotFoundError(f"Excel file not found: {full_path}")
    
    df = pd.read_excel(full_path, sheet_name=sheet_name, engine='openpyxl')
    return df


def read_excel_multiple_sheets(
    file_path: str,
    network_path: Optional[str] = None,
    auto_mount: bool = True
) -> dict[str, pd.DataFrame]:
    if auto_mount:
        try:
            mount_network(network_path)
        except Exception:
            pass
    
    full_path = get_mounted_path(file_path, network_path)
    
    if not os.path.exists(full_path):
        raise FileNotFoundError(f"Excel file not found: {full_path}")
    
    excel_file = pd.ExcelFile(full_path, engine='openpyxl')
    sheets_dict = {sheet_name: pd.read_excel(excel_file, sheet_name=sheet_name) 
                   for sheet_name in excel_file.sheet_names}
    return sheets_dict


def read_xlsx_columns_and_get_data(
    file_path: str,
    network_path: Optional[str] = None,
    auto_mount: bool = True,
    sheet_name: Optional[Union[str, int]] = 0
) -> pd.DataFrame:
    required_columns = ['CODIGO CEDENTE', 'GRUPO SACADO','CEDENTE', 'SACADO', 'PORTAL/EMAIL', 'LOGIN', 'SENHA']
    
    if auto_mount:
        try:
            mount_network(network_path)
        except Exception:
            pass
    
    full_path = get_mounted_path(file_path, network_path)
    
    if not os.path.exists(full_path):
        base_directory = os.path.dirname(full_path)
        if os.path.exists(base_directory):
            directory_items = os.listdir(base_directory)
            excel_files = [
                item for item in directory_items 
                if os.path.isfile(os.path.join(base_directory, item)) 
                and item.lower().endswith('.xlsx')
            ]
            
            if excel_files:
                first_excel_file = excel_files[0]
                base_path = os.path.dirname(file_path)
                file_path = os.path.join(base_path, first_excel_file)
            else:
                raise FileNotFoundError(f"Nenhum arquivo Excel encontrado em: {base_directory}")
        else:
            raise FileNotFoundError(f"Diretório não existe: {base_directory}")
    
    df = read_excel(file_path, sheet_name=sheet_name, network_path=network_path, auto_mount=False)
    
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Colunas obrigatórias não encontradas no arquivo: {', '.join(missing_columns)}")
    
    df_filtered = df[required_columns].copy()
    
    df_filtered = df_filtered.dropna(subset=['CEDENTE', 'GRUPO SACADO'], how='all')
    
    numeric_columns = ['CODIGO CEDENTE', 'GRUPO SACADO']
    for column in numeric_columns:
        if column in df_filtered.columns:
            df_filtered[column] = df_filtered[column].apply(
                lambda x: int(x) if pd.notna(x) else pd.NA
            )
    
    return df_filtered


def export_antecipations_to_excel(
    df_antecipations: pd.DataFrame,
    template_path: Optional[str] = None,
    output_path: Optional[str] = None,
    sheet_name: Optional[Union[str, int]] = 0
) -> str:
    if template_path is None:
        current_dir = Path(__file__).resolve().parents[1]
        template_path = current_dir / "template" / "template.xlsx"
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    if output_path is None:
        template_dir = os.path.dirname(template_path)
        timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M")
        output_path = os.path.join(template_dir, f"Titulos Concedidos para Antecipação_{timestamp}.xlsx")
    
    workbook = load_workbook(template_path)
    
    if isinstance(sheet_name, int):
        sheet = workbook.worksheets[sheet_name]
    else:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.worksheets[0]
    
    expected_columns = [
        'bordero', 'cedente', 'sacado', 'cnpj_sacado', 
        'titulo', 'valor', 'vencimento', 'portal/email', 'login', 'senha'
    ]
    
    df_export = df_antecipations.copy()
    
    if 'portal_email' in df_export.columns:
        df_export['portal/email'] = df_export['portal_email']
    
    columns_to_export = []
    for col in expected_columns:
        if col in df_export.columns:
            columns_to_export.append(col)
        else:
            df_export[col] = None
    
    df_export = df_export[expected_columns]
    
    header_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
        row_values = [cell.value for cell in row]
        if any(col in str(val).lower() if val else False for val in row_values for col in expected_columns):
            header_row = row_idx
            break
    
    if header_row is None:
        header_row = 1
    
    start_row = header_row + 1
    
    for row_idx, row_data in enumerate(dataframe_to_rows(df_export, index=False, header=False), start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if value is not None and not pd.isna(value):
                cell.value = value
    
    workbook.save(output_path)
    
    return output_path